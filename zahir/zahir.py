import os
import json
import sqlite3
import subprocess
import argparse
import glob
import gzip
from datetime import datetime
import openpyxl


#-----------------------------------history-section--------------------------------------- 

def extract_data_from_places(profile_path):
    db_path = f"{profile_path}/places.sqlite"
    if not os.path.exists(db_path):
        print(f"Places database not found: {db_path}")
        return None
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Extract website visits
    cursor.execute("""
        SELECT moz_places.url, moz_historyvisits.visit_date
        FROM moz_historyvisits
        JOIN moz_places ON moz_historyvisits.place_id = moz_places.id
    """)
    
    rows = cursor.fetchall()
    
    # Convert timestamp to readable date
    places_data = [
        {
            "url": row[0],
            "visit_date": datetime.fromtimestamp(row[1] / 1000000).strftime('%Y-%m-%d %H:%M:%S')
        } for row in rows
    ]
    
    conn.close()
    return places_data

def extract_data_from_logins(profile_path):
    logins_path = f"{profile_path}/logins.json"
    if not os.path.exists(logins_path):
        print(f"Logins file not found: {logins_path}")
        return None
    
    with open(logins_path, 'r') as file:
        logins_data = json.load(file)
    
    # Extract useful information
    logins_info = []
    for login in logins_data.get('logins', []):
        logins_info.append({
            "hostname": login.get('hostname', 'N/A'),
            "username": login.get('username', 'N/A'),
            "password": login.get('encryptedPassword', 'N/A')  # Password is encrypted
        })
    
    return logins_info

def extract_data_from_bookmarks(profile_path):
    db_path = f"{profile_path}/places.sqlite"
    if not os.path.exists(db_path):
        print(f"Places database not found: {db_path}")
        return None
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Extract bookmarks
    cursor.execute("""
        SELECT moz_places.url, moz_bookmarks.title
        FROM moz_bookmarks
        JOIN moz_places ON moz_bookmarks.fk = moz_places.id
    """)
    
    rows = cursor.fetchall()
    
    bookmarks_data = [
        {
            "url": row[0],
            "title": row[1]
        } for row in rows
    ]
    
    conn.close()
    return bookmarks_data

def combine_and_export_data(profile_path, output_dir):
    print("Extracting Firefox artifacts...")
    
    places_data = extract_data_from_places(profile_path)
    logins_data = extract_data_from_logins(profile_path)
    bookmarks_data = extract_data_from_bookmarks(profile_path)
    
    # Create a new workbook and add data to it
    wb = openpyxl.Workbook()
    
    # Places data (browser history)
    places_ws = wb.active
    places_ws.title = "Places (History)"
    places_ws.append(["URL", "Visit Date"])
    if places_data:
        for item in places_data:
            places_ws.append([item["url"], item["visit_date"]])
    
    # Logins data
    logins_ws = wb.create_sheet(title="Logins")
    logins_ws.append(["Hostname", "Username", "Password (Encrypted)"])
    if logins_data:
        for item in logins_data:
            logins_ws.append([item["hostname"], item["username"], item["password"]])
    
    # Bookmarks data
    bookmarks_ws = wb.create_sheet(title="Bookmarks")
    bookmarks_ws.append(["URL", "Title"])
    if bookmarks_data:
        for item in bookmarks_data:
            bookmarks_ws.append([item["url"], item["title"]])
    
    # Save the workbook
    output_file = f"{output_dir}/firefox_artifacts.xlsx"
    wb.save(output_file)
    
    print(f"Data extraction complete. Results saved to {output_file}.")
    
#-----------------------------------------recovery-section------------------------------------

def recover_deleted_files_from_usb(mount_point, output_dir, device):
    try:
        # Check if the device is mounted
        mount_check = subprocess.run(["mount"], capture_output=True, text=True)
        if device in mount_check.stdout:
            # Unmount the device if it is mounted
            subprocess.run(["sudo", "umount", device], check=True)
            print(f"Device {device} unmounted successfully.")
        else:
            print(f"Device {device} is not mounted, proceeding with recovery.")
        
        # Scan the device to list deleted files
        print("Scanning for deleted files...")
        scan_result = subprocess.run(["sudo", "ntfsundelete", device, "--scan"], check=True, capture_output=True, text=True)
        print(scan_result.stdout)  # Show the scan result to the user

        # Prompt the user for the inode of the file(s) to recover
        inodes = input("Enter the inode(s) of the file(s) to recover, separated by commas: ")
        if not inodes:
            raise Exception("No inodes provided, unable to recover files.")
        
        # Recover the specified inodes
        command = ["sudo", "ntfsundelete", device, "--undelete", "--inodes", inodes, "--destination", output_dir]
        subprocess.run(command, check=True)
        print(f"Deleted files recovered successfully. Check {output_dir} for recovered files.")
        
    except subprocess.CalledProcessError as e:
        print(f"Error during recovery: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")



#--------------------------------------------logs-section-------------------------------------
def collect_and_analyze_logs(output_dir):
    print("Collecting and analyzing system and application logs...")

    # Define the list of log files to collect
    log_files = glob.glob("/var/log/*.log")

    logs_output = {}
    log_analysis_report = {}

    # Ensure output directory exists
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Collect logs
    for log_file in log_files:
        if os.path.exists(log_file):
            try:
                # Read the log file content
                with open(log_file, 'r', errors='ignore') as f:
                    content = f.readlines()
                    logs_output[log_file] = content  # Store log content in memory
                    log_analysis_report[log_file] = analyze_log(content)  # Analyze the log content

            except Exception as e:
                print(f"Error processing {log_file}: {e}")

    # Save the collected logs and analysis report in Excel format
    save_logs_and_report(output_dir, logs_output, log_analysis_report)


def analyze_log(log_content):
    """
    Analyzes the log content for specified keywords.
    """
    # Define keywords for analysis
    keywords = [
        'error', 'failed', 'denied', 'segfault', 'unauthorized', 'critical', 'panic', 'fatal', 'exception', 'corruption',
        'abort', 'failure', 'unreachable', 'timeout', 'overload', 'missing', 'halt', 'invalid', 'forbidden', 'unauthenticated',
        'unauthorized access', 'illegal', 'intrusion', 'violation', 'suspicious', 'compromised', 'exploited', 'breach', 'malicious',
        'tampered', 'cipher', 'certificate', 'disconnected', 'connection lost', 'dropped', 'broken pipe', 'network unreachable',
        'reset', 'unresponsive', 'unavailable', 'refused', 'disk full', 'read-only', 'out of space', 'corrupted', 'filesystem',
        'i/o error', 'mount failure', 'device not found', 'quota exceeded', 'inconsistent', 'crash', 'service down', 'daemon',
        'zombie process', 'core dumped', 'service failure', 'terminated', 'restarting', 'memory leak', 'out of memory', 'cpu overload',
        'gpu failure', 'hardware failure', 'overheated', 'throttling', 'fan failure', 'temperature', 'trojan', 'virus', 'exploit',
        'backdoor', 'spyware', 'ransomware', 'rootkit', 'malware', 'phishing', 'payload', 'infected', 'deprecated', 'warning', 'notice',
        'alert', 'shutdown', 'reboot', 'restart', 'signal', 'upgrade required', 'lockout', 'expiration', 'high latency', 'deadlock',
        'lock wait timeout', 'index corruption', 'transaction failure', 'query timeout', 'rollback', 'failed commit', 'read error',
        'write error'
    ]
    analysis = {'lines': [], 'keywords_found': []}

    for line in log_content:
        for keyword in keywords:
            if keyword in line.lower():
                analysis['lines'].append(line.strip())
                analysis['keywords_found'].append(keyword)

    return analysis


def save_logs_and_report(output_dir, logs_output, log_analysis_report):
    """
    Saves the collected logs and analysis reports to an Excel file.
    """
    # Create a new workbook
    wb = openpyxl.Workbook()

    # Collected Logs
    logs_ws = wb.active
    logs_ws.title = "Collected Logs"
    logs_ws.append(["Log File", "Log Content"])

    for log_file, log_content in logs_output.items():
        log_content_str = "\n".join(log_content)  # Combine log lines into a single string
        logs_ws.append([log_file, log_content_str])

    # Log Analysis Report
    analysis_ws = wb.create_sheet(title="Log Analysis Report")
    analysis_ws.append(["Log File", "Keywords Found", "Relevant Log Lines"])

    for log_file, analysis in log_analysis_report.items():
        keywords_str = ", ".join(analysis['keywords_found']) if analysis['keywords_found'] else "None"
        lines_str = "\n".join(analysis['lines']) if analysis['lines'] else "None"
        analysis_ws.append([log_file, keywords_str, lines_str])

    # Save the workbook
    output_file = os.path.join(output_dir, "logs_and_analysis_report.xlsx")
    wb.save(output_file)

    print(f"Log files and analysis report saved to {output_file}.")

#-----------------------------------user-activity-section------------------------------------

def reconstruct_user_activity(output_dir, username):
    print(f"Reconstructing user activity for {username}...")

    user_activity = {
        'command_history': [],
        'file_access_history': [],
        'recently_accessed_files': [],
        'login_logout_history': []
    }

    # Adjust paths based on the provided username
    history_file = f"/home/{username}/.bash_history"  # Path for bash history of provided user
    audit_log_file = "/var/log/audit/audit.log"  # Path for audit logs
    last_log_cmd = "last"  # Command to get login/logout history
    recent_files_cmd = f"lsof -u {username} 2>/dev/null"  # Command to get recently accessed files

    # Ensure output directory exists
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Collect command history from /home/{username}/.bash_history
    if os.path.exists(history_file):
        try:
            result = subprocess.check_output(f"sudo cat {history_file}", shell=True, stderr=subprocess.STDOUT)
            commands = result.decode('utf-8').splitlines()
            user_activity['command_history'] = [cmd.strip() for cmd in commands]
        except subprocess.CalledProcessError as e:
            print(f"Error reading {history_file} with sudo: {e.output.decode('utf-8')}")
        except Exception as e:
            print(f"Error reading {history_file}: {e}")
    else:
        print(f"{history_file} does not exist or is inaccessible.")

    # Collect file access history from /var/log/audit/audit.log
    if os.path.exists(audit_log_file):
        try:
            with open(audit_log_file, 'r', errors='ignore') as f:
                file_access = f.readlines()
                user_activity['file_access_history'] = [line.strip() for line in file_access if 'type=PATH' in line]
        except Exception as e:
            print(f"Error reading {audit_log_file}: {e}")
    else:
        print(f"{audit_log_file} does not exist or is inaccessible. Ensure auditd is enabled and logging.")

    # Collect login/logout history using 'last' command
    try:
        result = subprocess.check_output(last_log_cmd, shell=True).decode('utf-8')
        user_activity['login_logout_history'] = result.splitlines()
    except subprocess.CalledProcessError as e:
        print(f"Error executing '{last_log_cmd}' command: {e}")
        print("Please ensure 'last' is installed: sudo apt-get install util-linux")

    # Collect recently accessed files using 'lsof'
    try:
        result = subprocess.check_output(recent_files_cmd, shell=True).decode('utf-8')
        user_activity['recently_accessed_files'] = result.splitlines()
    except subprocess.CalledProcessError as e:
        print(f"Error executing '{recent_files_cmd}' command: {e}")

    # Save user activity data to spreadsheet
    save_to_spreadsheet(user_activity, output_dir)

    print("User activity reconstruction completed.")

def save_to_spreadsheet(user_activity, output_dir):
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()

    # Create sheets for each type of activity
    sheets = {
        'Command History': user_activity['command_history'],
        'File Access History': user_activity['file_access_history'],
        'Recently Accessed Files': user_activity['recently_accessed_files'],
        'Login_Logout History': user_activity['login_logout_history']
    }

    for sheet_name, data in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for idx, entry in enumerate(data, start=1):
            ws.cell(row=idx, column=1, value=entry)

    # Remove the default empty sheet created by openpyxl
    del wb['Sheet']

    # Define the output file path
    activity_file_path = os.path.join(output_dir, "user_activity.xlsx")

    # Save the workbook
    try:
        wb.save(activity_file_path)
        print(f"User activity data saved to {activity_file_path}.")
    except Exception as e:
        print(f"Error saving user activity data to spreadsheet: {e}")

        
